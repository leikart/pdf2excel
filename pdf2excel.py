"""
PDF 轉 Excel 通用工具 v4
- 雙模式：自動表格偵測 / 文字座標精準解析
- 座標模式：自動從標題列計算欄界、合併同列多欄內容
- 浮水印過濾、重複標題刪除、錨點欄續行合併
- 合併儲存格（出貨日期/星期同天合併）
- 預覽 + 欄位名稱調整 + 另存選項
- 批次轉換
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, re
from openpyxl import load_workbook as _load_wb

try:
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable,"-m","pip","install","pdfplumber","openpyxl","-q"])
    import pdfplumber
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

BG="#12121F"; BG2="#1A1A2E"; CARD="#0D2137"; ACCENT="#E94560"
ACCENT2="#6C3FC4"; TEXT="#F0F0F0"; TEXT2="#8FA8C8"
SUCCESS="#3EC97A"; ERROR="#E94560"; WARNING="#F5A623"
ACCENT_HOVER="#FF6B81"; CARD2="#162840"

DEFAULT_SETTINGS = {
    "mode": "auto",
    "watermark_keywords": "",
    "anchor_col": "0",
    "skip_duplicate_header": True,
    "col_bounds": "",
    "merge_col_range": "",
}

def clean(v):
    if v is None: return ""
    return re.sub(r"\s+"," ",str(v)).strip()

def rows_equal(a, b):
    ca=[clean(x) for x in (a or [])]
    cb=[clean(x) for x in (b or [])]
    return ca==cb and any(ca)

def normalize(table):
    return [[clean(v) for v in row] for row in table if any(clean(v) for v in row)]

# ══════════════════════════════════════════
#  模式一：自動表格偵測
# ══════════════════════════════════════════
T_LINES={"vertical_strategy":"lines","horizontal_strategy":"lines",
         "snap_tolerance":3,"join_tolerance":3,"edge_min_length":3,
         "min_words_vertical":1,"min_words_horizontal":1,
         "intersection_tolerance":3,"text_tolerance":3}
T_TEXT={**T_LINES,"vertical_strategy":"text","horizontal_strategy":"text"}

def _extract_page_mixed(page, wm_keys, skip_dup, header, all_rows):
    """
    混合模式頁面解析：
    - 有框線的部分：用 cell bbox 擷取（保留換行）
    - 框線外的文字（說明文字/標題）：依 y 座標插入第一欄
    回傳 (updated_header, strategy)
    """
    all_words = page.extract_words(x_tolerance=3, y_tolerance=3)
    line_map = {}
    for w in all_words:
        y = round(w["top"]/3)*3
        line_map.setdefault(y, []).append(w)

    T_LINES_LOCAL={"vertical_strategy":"lines","horizontal_strategy":"lines",
                   "snap_tolerance":3,"join_tolerance":3,"edge_min_length":3,
                   "min_words_vertical":1,"min_words_horizontal":1,
                   "intersection_tolerance":3,"text_tolerance":3}

    tbl_objs = page.find_tables(T_LINES_LOCAL)
    strategy = "lines+text" if tbl_objs else "text"

    def _flat(row): return " ".join(v.replace("\n"," ") if v else "" for v in row)
    def _is_wm(row): return any(k in _flat(row) for k in wm_keys) if wm_keys else False
    def _is_hdr(row):
        t = _flat(row)
        return ("出貨單編號" in t and "出貨日期" in t) or ("出貨日期" in t and "星期" in t and "位置" in t)

    page_rows = []  # (y, row_data)

    if tbl_objs:
        tbl_bboxes = [t.bbox for t in tbl_objs]

        # ── 表格內容（cell level，保留換行）──
        for tbl in tbl_objs:
            for tbl_row in tbl.rows:
                row_data = []; row_y = None
                for cell in tbl_row.cells:
                    if cell is None: row_data.append(""); continue
                    x0,y0,x1,y1 = cell
                    if row_y is None: row_y = (y0+y1)/2
                    cell_lines = {}
                    for y, ws in line_map.items():
                        for w in ws:
                            if y0<=w["top"]<=y1 and x0<=w["x0"]<=x1:
                                cell_lines.setdefault(round(w["top"],1),[]).append(w["text"])
                    text = "\n".join(" ".join(cell_lines[y]) for y in sorted(cell_lines))
                    row_data.append(text)
                if any(row_data) and row_y is not None:
                    page_rows.append((row_y, row_data))

        # ── 表格外的文字（說明文字，放入第一欄）──
        for y_key in sorted(line_map.keys()):
            ws_line = sorted(line_map[y_key], key=lambda w: w["x0"])
            if not ws_line: continue
            actual_y = ws_line[0]["top"]
            line_text = " ".join(w["text"] for w in ws_line)
            if not line_text.strip(): continue
            in_tbl = any(ty0 <= actual_y <= ty1 for (_,ty0,_,ty1) in tbl_bboxes)
            if not in_tbl:
                page_rows.append((actual_y, [line_text]))
    else:
        # 純文字模式（無框線）
        T_TEXT_LOCAL={**T_LINES_LOCAL,"vertical_strategy":"text","horizontal_strategy":"text"}
        tables = page.extract_tables(T_TEXT_LOCAL)
        if tables:
            strategy = "text"
            for t in tables:
                for row in t:
                    r = [clean(v) for v in row]
                    if any(r):
                        page_rows.append((0, r))
        else:
            for y_key in sorted(line_map.keys()):
                ws_line = sorted(line_map[y_key], key=lambda w: w["x0"])
                if not ws_line: continue
                actual_y = ws_line[0]["top"]
                line_text = " ".join(w["text"] for w in ws_line)
                if line_text.strip():
                    page_rows.append((actual_y, [line_text]))

    # 依 y 排序
    page_rows.sort(key=lambda x: x[0])

    # 過濾浮水印、處理標題列、加入結果
    for _, row_data in page_rows:
        if _is_wm(row_data): continue
        if _is_hdr(row_data):
            if header is None:
                header = [clean(v) for v in row_data]
                all_rows.append(header)
            elif not skip_dup:
                all_rows.append([clean(v) for v in row_data])
            continue
        all_rows.append(row_data)

    return header, strategy

def parse_auto(pdf_path, settings, log_cb=None):
    wm=[k.strip() for k in settings["watermark_keywords"].split(",") if k.strip()]
    skip_dup=settings.get("skip_duplicate_header",True)
    all_rows=[]; header=None; strategy="lines"

    with pdfplumber.open(pdf_path) as pdf:
        if log_cb: log_cb(f"  共 {len(pdf.pages)} 頁（混合模式）")
        for pn,page in enumerate(pdf.pages,1):
            if log_cb: log_cb(f"  解析第 {pn} 頁...")
            header, page_strategy = _extract_page_mixed(page, wm, skip_dup, header, all_rows)
            strategy = page_strategy

    if settings.get("fill_dates", True):
        all_rows = _fill_date_cols(all_rows)
    return all_rows, strategy


def _fill_date_cols(rows):
    """
    把日期/星期欄的空白格填入上一列的值（向下填充）。
    只對「錨點欄（第0欄）有值」的列填充，避免把日期錯誤填進合計列或說明文字列。
    """
    if len(rows) < 2: return rows
    header = rows[0]
    # 找日期欄和星期欄的索引
    date_cols = []
    for ci, h in enumerate(header):
        if any(k in str(h) for k in ["日期","星期","週","week","date"]):
            date_cols.append(ci)
    if not date_cols:
        return rows  # 找不到日期/星期欄，不做填充直接回傳
    last_vals = {ci: "" for ci in date_cols}
    result = [header]
    for row in rows[1:]:
        new_row = list(row)
        # 錨點欄（第0欄）的值
        anchor_val = str(new_row[0]).strip() if new_row else ""
        for ci in date_cols:
            val = new_row[ci] if ci < len(new_row) else ""
            val_str = str(val).strip() if val else ""
            if val_str:
                # 有值：更新記憶
                last_vals[ci] = val_str
            elif last_vals[ci] and anchor_val:
                # 空白 + 錨點欄有值 → 填入上一列的值
                while len(new_row) <= ci: new_row.append("")
                new_row[ci] = last_vals[ci]
            # 錨點欄空白（合計列、說明文字）→ 不填充，保持空白
        result.append(new_row)
    return result

# ══════════════════════════════════════════
#  模式二：文字座標精準解析
# ══════════════════════════════════════════
def _calc_bounds_from_header(header_words):
    """從標題列文字位置自動計算欄界（取相鄰欄的中間點）"""
    sorted_h = sorted(header_words, key=lambda w: w["x0"])
    bounds = [0.0]
    for i in range(len(sorted_h)-1):
        mid = (sorted_h[i]["x1"] + sorted_h[i+1]["x0"]) / 2
        bounds.append(round(mid, 1))
    bounds.append(9999.0)
    return bounds

def _get_col(x, bounds):
    for i in range(len(bounds)-1):
        if bounds[i] <= x < bounds[i+1]: return i
    return len(bounds)-2

def _is_header(line_text, hdr_keys):
    return sum(1 for k in hdr_keys if k in line_text) >= 2

def _is_watermark(line_text, wm_keys):
    return any(k in line_text for k in wm_keys)

def parse_coords(pdf_path, settings, log_cb=None):
    wm  = [k.strip() for k in settings["watermark_keywords"].split(",") if k.strip()]
    skip_dup = settings.get("skip_duplicate_header", True)
    anchor   = int(settings.get("anchor_col","0") or 0)

    # 手動欄界
    cb_str = settings.get("col_bounds","").strip()
    manual_bounds = None
    if cb_str:
        try:
            nums = [float(x) for x in cb_str.split(",") if x.strip()]
            manual_bounds = sorted(nums) + [9999.0]
        except: pass

    all_rows=[]; bounds=manual_bounds; hdr_keys=[]; col_count=0; header_row=None

    with pdfplumber.open(pdf_path) as pdf:
        if log_cb: log_cb(f"  共 {len(pdf.pages)} 頁（座標模式）")
        for pn,page in enumerate(pdf.pages,1):
            if log_cb: log_cb(f"  解析第 {pn} 頁...")
            words = page.extract_words(x_tolerance=3,y_tolerance=3,
                                       keep_blank_chars=False,use_text_flow=False)
            if not words: continue

            # 依 y 分組
            lines={}
            for w in words:
                y=round(w["top"]/4)*4
                lines.setdefault(y,[]).append(w)

            for y in sorted(lines.keys()):
                ws=sorted(lines[y],key=lambda w:w["x0"])
                lt=" ".join(w["text"] for w in ws)

                # 過濾浮水印
                if wm and _is_watermark(lt, wm): continue

                # 偵測標題列：自動計算欄界
                if bounds is None or (_is_header(lt, hdr_keys if hdr_keys else ["出貨","日期","位置","編號"])):
                    if bounds is None:
                        bounds = _calc_bounds_from_header(ws)
                        col_count = len(bounds)-1
                        hdr_keys  = [w["text"] for w in ws]
                        if log_cb: log_cb(f"  偵測到 {col_count} 欄，欄界：{[round(b) for b in bounds[:-1]]}")
                    # 標題列：第一次加入，之後跳過（skip_dup）
                    if header_row is None:
                        header_row = [w["text"] for w in ws]
                        # 展開成欄位陣列
                        row=[""]*(col_count or len(ws))
                        for w in ws:
                            ci=_get_col(w["x0"],bounds)
                            if ci<len(row): row[ci]=(row[ci]+" "+w["text"]).strip()
                        all_rows.append(row)
                    elif not skip_dup:
                        row=[""]*(col_count)
                        for w in ws:
                            ci=_get_col(w["x0"],bounds)
                            if ci<len(row): row[ci]=(row[ci]+" "+w["text"]).strip()
                        all_rows.append(row)
                    continue

                if bounds is None: continue

                # 一般資料列
                row=[""]*(col_count or 8)
                for w in ws:
                    ci=_get_col(w["x0"],bounds)
                    if ci<len(row): row[ci]=(row[ci]+" "+w["text"]).strip()
                if any(row): all_rows.append(row)

    # 合併指定欄（如把「星期欄」和「位置欄」分開時位置被切兩半）
    merge_range = settings.get("merge_col_range","").strip()
    if merge_range:
        try:
            parts=merge_range.split("-")
            mc_start=int(parts[0]); mc_end=int(parts[1])
            merged_rows=[]
            for row in all_rows:
                new_row=list(row[:mc_start])
                combined=" ".join(row[ci] for ci in range(mc_start,min(mc_end+1,len(row))) if row[ci])
                new_row.append(combined)
                new_row.extend(row[mc_end+1:])
                merged_rows.append(new_row)
            all_rows=merged_rows
        except: pass

    # 處理錨點欄續行
    result=_merge_continuation(all_rows, anchor)
    return result, "coords"

def _merge_continuation(rows, anchor_col):
    """錨點欄空白的列 → 合併到上一筆的備註欄"""
    if not rows: return []
    result=[]
    for row in rows:
        av=row[anchor_col] if anchor_col<len(row) else ""
        # 判斷是否為標題列
        is_hdr=any(k in " ".join(row) for k in ["出貨單編號","出貨日期","位置","數量","進場時間"])
        if is_hdr and not result:
            result.append(row); continue
        if not av and result and not is_hdr:
            extra=" ".join(v for v in row if v)
            if extra:
                last=result[-1]
                note=last[-1] if last else ""
                result[-1]=last[:-1]+[((note+"\n"+extra).strip() if note else extra)]
        else:
            result.append(row)
    return result

def _merge_tables(all_tables, settings):
    skip=settings.get("skip_duplicate_header",True)
    merged=[]; header=None
    for t in all_tables:
        if not t: continue
        if header is None:
            header=t[0]; merged.extend(t)
        else:
            start=1 if (skip and rows_equal(t[0],header)) else 0
            merged.extend(t[start:])
    return merged

# ══════════════════════════════════════════
#  合併儲存格偵測
# ══════════════════════════════════════════
def detect_merges(table, merge_cols=None):
    """
    偵測需要合併的欄位範圍。
    只合併標題列中含有「日期」「星期」等關鍵字的欄位。
    找不到這類欄位時回傳空字典（不合併任何欄位）。
    merge_cols: 可手動指定欄位索引清單，覆蓋自動偵測結果。
    """
    if len(table)<2: return {}
    data=table[1:]; info={}

    if merge_cols is None:
        # 從標題列找日期/星期欄，找不到就不合併
        header=table[0]
        date_cols=[]
        for ci,h in enumerate(header):
            if any(k in str(h) for k in ["日期","星期","週","week","date","Date"]):
                date_cols.append(ci)
        if not date_cols:
            return {}  # 沒有日期欄 → 不合併任何欄位
        merge_cols=date_cols

    for ci in merge_cols:
        ranges=[]; i=0
        while i<len(data):
            val=data[i][ci] if ci<len(data[i]) else ""
            if val:
                j=i+1
                while j<len(data):
                    nv=data[j][ci] if ci<len(data[j]) else ""
                    if nv=="": j+=1   # 空白代表同一組，繼續合併
                    else: break        # 有新值代表換組，停止
                if j-i>1: ranges.append((i,j-1))
                i=j
            else: i+=1
        if ranges: info[ci]=ranges
    return info

# ══════════════════════════════════════════
#  寫入 Excel
# ══════════════════════════════════════════
def write_excel(table, merge_info, out_path, col_map=None):
    wb=Workbook(); ws=wb.active; ws.title="轉換結果"
    hfill=PatternFill(start_color="0F3460",end_color="0F3460",fill_type="solid")
    hfont=Font(color="FFFFFF",bold=True,size=11)
    mfill=PatternFill(start_color="D6E4F0",end_color="D6E4F0",fill_type="solid")
    o_fill=PatternFill(start_color="F5F9FF",end_color="F5F9FF",fill_type="solid")
    e_fill=PatternFill(start_color="FFFFFF",end_color="FFFFFF",fill_type="solid")
    thin=Side(style="thin")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="top",wrap_text=True)
    lwrap=Alignment(horizontal="left",vertical="top",wrap_text=True)
    if not table: wb.save(out_path); return
    ncols=max(len(r) for r in table)

    # 欄寬：取各欄最長內容（換行只算最長的一段），上限50
    col_max=[12]*ncols
    for row in table:
        for ci,v in enumerate(row):
            if v:
                # 換行內容取最長一行來計算欄寬
                max_line=max((len(line) for line in str(v).split("\n")),default=0)
                col_max[ci]=min(max(col_max[ci], max_line), 50)
    for ci in range(ncols):
        ws.column_dimensions[get_column_letter(ci+1)].width=col_max[ci]+3

    drc=0
    for ri,row in enumerate(table):
        er=ri+1; is_hdr=(ri==0)
        max_lines=1  # 這一列最多幾行
        for ci in range(ncols):
            val=row[ci] if ci<len(row) else ""
            if is_hdr and col_map and ci in col_map: val=col_map[ci]
            # 保留換行符（\n → Excel 換行）
            cell=ws.cell(row=er,column=ci+1,value=val)
            cell.border=bdr
            cell.alignment=ctr if is_hdr else lwrap
            if is_hdr:
                cell.fill=hfill; cell.font=hfont
            else:
                cell.fill=o_fill if drc%2==0 else e_fill
                cell.font=Font(size=10)
                if val: max_lines=max(max_lines, str(val).count("\n")+1)
        if not is_hdr:
            # 列高：依各欄換行數計算，每行15pt，最少22，最多300
            ws.row_dimensions[er].height=min(max(22, 15*max_lines), 300)
            drc+=1
        else:
            ws.row_dimensions[er].height=24
    ws.freeze_panes="A2"
    for ci,ranges in merge_info.items():
        for (si,ei) in ranges:
            sr=si+2; er2=ei+2
            if sr==er2: continue
            try:
                ws.merge_cells(start_row=sr,start_column=ci+1,end_row=er2,end_column=ci+1)
                cell=ws.cell(row=sr,column=ci+1)
                cell.alignment=ctr; cell.fill=mfill
                cell.font=Font(size=10,bold=True,color="1F3864"); cell.border=bdr
            except: pass
    wb.save(out_path)

def convert(pdf_path, out_path, settings, log_cb=None):
    mode=settings.get("mode","auto")
    if mode=="coords":
        table,strategy=parse_coords(pdf_path,settings,log_cb)
    elif mode=="table":
        table,strategy=parse_auto(pdf_path,settings,log_cb)
    else:
        table,strategy=parse_auto(pdf_path,settings,log_cb)
        if not table or len(table)<3:
            if log_cb: log_cb("  表格偵測結果少，切換座標模式...")
            table,strategy=parse_coords(pdf_path,settings,log_cb)
    if log_cb: log_cb(f"  策略：{strategy}，{len(table)} 列（含標題）")
    # 依文件類型決定是否合併儲存格
    merge_dates=settings.get("merge_dates", True)
    mi=detect_merges(table) if merge_dates else {}
    if out_path: write_excel(table,mi,out_path)
    return table, strategy


# ════════════════════════════════════════════════════════
#  比對核心
# ════════════════════════════════════════════════════════

def load_table_from_file(path, settings, log_cb=None):
    ext=os.path.splitext(path)[1].lower()
    if ext==".pdf":
        if log_cb: log_cb(f"  轉換 PDF：{os.path.basename(path)}")
        table,_=convert(path, None, settings, log_cb=log_cb)
        return table
    elif ext in (".xlsx",".xls"):
        if log_cb: log_cb(f"  載入 Excel：{os.path.basename(path)}")
        wb=_load_wb(path, data_only=True)
        ws=wb.active
        rows=[]
        for row in ws.iter_rows(values_only=True):
            r=[clean(v) for v in row]
            if any(r): rows.append(r)
        return rows
    else:
        raise ValueError(f"不支援的格式：{ext}")

def compare_tables(table_a, table_b, compare_cols):
    data_a=table_a[1:] if len(table_a)>1 else []
    data_b=table_b[1:] if len(table_b)>1 else []
    header=table_a[0] if table_a else []
    max_len=max(len(data_a),len(data_b))
    results=[]
    for i in range(max_len):
        row_a=data_a[i] if i<len(data_a) else None
        row_b=data_b[i] if i<len(data_b) else None
        if row_a is None:
            results.append({"type":"add","row_a":None,"row_b":row_b,"diffs":[],"idx":i})
        elif row_b is None:
            results.append({"type":"delete","row_a":row_a,"row_b":None,"diffs":[],"idx":i})
        else:
            diffs=[]
            for ci in compare_cols:
                va=row_a[ci] if ci<len(row_a) else ""
                vb=row_b[ci] if ci<len(row_b) else ""
                if clean(va)!=clean(vb):
                    col_name=header[ci] if ci<len(header) else f"欄{ci+1}"
                    diffs.append((ci,col_name,clean(va),clean(vb)))
            t="modify" if diffs else "match"
            results.append({"type":t,"row_a":row_a,"row_b":row_b,"diffs":diffs,"idx":i})
    return results, header

def write_diff_excel(results, header, compare_cols, out_path):
    wb=Workbook(); ws=wb.active; ws.title="差異報告"
    hfill=PatternFill(start_color="0F3460",end_color="0F3460",fill_type="solid")
    hfont=Font(color="FFFFFF",bold=True,size=11)
    fills={"add":PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid"),
           "delete":PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid"),
           "modify":PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")}
    thin=Side(style="thin")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="top",wrap_text=True)
    lwrap=Alignment(horizontal="left",vertical="top",wrap_text=True)
    col_names_a=[f"{header[ci] if ci<len(header) else f'欄{ci+1}'}\n（檔案A）" for ci in compare_cols]
    col_names_b=[f"{header[ci] if ci<len(header) else f'欄{ci+1}'}\n（檔案B）" for ci in compare_cols]
    hdr_row=["差異類型","列號"]+col_names_a+col_names_b+["差異說明"]
    for ci,h in enumerate(hdr_row,1):
        c=ws.cell(row=1,column=ci,value=h)
        c.fill=hfill; c.font=hfont; c.border=bdr; c.alignment=ctr
    ws.row_dimensions[1].height=30
    ri=2
    type_map={"add":"新增","delete":"刪除","modify":"修改"}
    for res in results:
        t=res["type"]
        if t=="match": continue
        row_a=res["row_a"] or []; row_b=res["row_b"] or []
        idx=res["idx"]+2
        if t=="add": desc=f"第{idx}列：檔案B新增此列"
        elif t=="delete": desc=f"第{idx}列：此列在檔案B中已刪除"
        else:
            parts=[f"【{cn}】A=「{va}」→ B=「{vb}」" for _,cn,va,vb in res["diffs"]]
            desc=f"第{idx}列修改：\n"+"\n".join(parts)
        vals_a=[clean(row_a[ci]) if ci<len(row_a) else "" for ci in compare_cols]
        vals_b=[clean(row_b[ci]) if ci<len(row_b) else "" for ci in compare_cols]
        row_data=[type_map[t],idx]+vals_a+vals_b+[desc]
        nl=desc.count("\n")+1
        ws.row_dimensions[ri].height=max(18,15*nl)
        for ci,val in enumerate(row_data,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.fill=fills.get(t,fills["modify"]); c.border=bdr
            c.alignment=ctr if ci<=2 else lwrap; c.font=Font(size=10)
        ri+=1
    if ri==2:
        c=ws.cell(row=2,column=1,value="✅ 兩份檔案在選擇的欄位中完全相同")
        c.font=Font(size=12,bold=True,color="375623")
    ws.column_dimensions["A"].width=10
    ws.column_dimensions["B"].width=8
    ncol=len(compare_cols)
    for i in range(ncol*2):
        ws.column_dimensions[get_column_letter(3+i)].width=24
    ws.column_dimensions[get_column_letter(3+ncol*2)].width=55
    ws.freeze_panes="A2"
    wb.save(out_path)

# ══════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF 轉 Excel 通用工具 v4")
        self.configure(bg=BG)
        self.after(0, lambda: self.state("zoomed"))
        self.minsize(1000,700)
        self.pdf_files=[]; self.running=False
        self.settings=dict(DEFAULT_SETTINGS)
        self.out_dir=tk.StringVar(value="")
        s=ttk.Style(); s.theme_use("clam")
        s.configure("P.Horizontal.TProgressbar",troughcolor=CARD,background=ACCENT,
                    darkcolor=ACCENT,lightcolor=ACCENT,bordercolor=BG,thickness=14)
        s.configure("Treeview",background=CARD,foreground=TEXT,
                    fieldbackground=CARD,font=("Segoe UI",11),rowheight=26)
        s.configure("Treeview.Heading",background=BG2,foreground=TEXT,
                    font=("Segoe UI",11,"bold"))
        # 分頁 Notebook 樣式
        s.configure("App.TNotebook",background=BG,borderwidth=0,tabmargins=0)
        s.configure("App.TNotebook.Tab",
                    background=BG2,foreground=TEXT2,
                    font=("Segoe UI",12,"bold"),
                    padding=(24,10),borderwidth=0)
        s.map("App.TNotebook.Tab",
              background=[("selected",CARD),("active","#1A3050")],
              foreground=[("selected","white"),("active",TEXT)])
        self._build()

    # ── 輔助：分隔線 ─────────────────────────────
    def _sep(self, parent, color="#2A3A50"):
        tk.Frame(parent,bg=color,height=1).pack(fill="x",pady=6)

    # ── 輔助：區塊標題 ───────────────────────────
    def _section(self, parent, text):
        tk.Label(parent,text=text,font=("Segoe UI",12,"bold"),
                 bg=BG,fg=TEXT2).pack(anchor="w",pady=10)

    # ── 輔助：帶底色的卡片 Frame ─────────────────
    def _card(self, parent, bg=None, pad=10):
        f=tk.Frame(parent,bg=bg or CARD,padx=pad,pady=pad)
        f.pack(fill="x",pady=(0,6))
        return f

    def _build(self):
        # ══ 頂部標題列 ════════════════════════════
        hdr=tk.Frame(self,bg=CARD,height=64); hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr,text="⬛",font=("Segoe UI",18),bg=ACCENT,fg="white",
                 width=3).pack(side="left",fill="y")
        tk.Label(hdr,text=" PDF  →  Excel",font=("Segoe UI",20,"bold"),
                 bg=CARD,fg="white").pack(side="left",padx=(8,0),pady=14)
        tk.Label(hdr,text="v4",font=("Segoe UI",9),bg="#1A3050",fg=TEXT2,
                 padx=8).pack(side="right",padx=14,pady=18)
        tk.Label(hdr,text="通用版・混合模式解析",font=("Segoe UI",10),
                 bg=CARD,fg=TEXT2).pack(side="right",padx=0,pady=18)

        # ══ 分頁 Notebook ═════════════════════════
        self.nb=ttk.Notebook(self,style="App.TNotebook")
        self.nb.pack(fill="both",expand=True,padx=0,pady=0)

        # 頁籤1：轉換
        self.tab_convert=tk.Frame(self.nb,bg=BG)
        self.nb.add(self.tab_convert,text="  📄  PDF 轉 Excel  ")

        # 頁籤2：比對
        self.tab_compare=tk.Frame(self.nb,bg=BG)
        self.nb.add(self.tab_compare,text="  🔍  檔案比對  ")

        self._build_convert_tab(self.tab_convert)
        self._build_compare_tab(self.tab_compare)

    def _build_convert_tab(self, parent):
        main=tk.Frame(parent,bg=BG); main.pack(fill="both",expand=True,padx=14,pady=10)

        # ══ 左側：PDF 清單 ════════════════════════
        left=tk.Frame(main,bg=BG); left.pack(side="left",fill="both",expand=True)

        r1=tk.Frame(left,bg=BG); r1.pack(fill="x",pady=6)
        tk.Label(r1,text="PDF 檔案",font=("Segoe UI",12,"bold"),bg=BG,fg=TEXT).pack(side="left")
        tk.Button(r1,text="  ＋  新增檔案  ",font=("Segoe UI",10,"bold"),
                  bg=ACCENT,fg="white",relief="flat",cursor="hand2",pady=5,
                  command=self.add_files).pack(side="right",padx=(4,0))
        tk.Button(r1,text="清除",font=("Segoe UI",10),
                  bg=BG2,fg=TEXT2,relief="flat",cursor="hand2",padx=12,pady=5,
                  command=self.clear_files).pack(side="right")

        # 檔案清單（帶圓角外框視覺）
        lf=tk.Frame(left,bg="#0A1828",bd=0); lf.pack(fill="both",expand=True)
        sb=tk.Scrollbar(lf); sb.pack(side="right",fill="y")
        self.listbox=tk.Listbox(lf,yscrollcommand=sb.set,
                                bg="#0A1828",fg=TEXT,selectbackground=ACCENT,
                                selectforeground="white",
                                font=("Segoe UI",11),relief="flat",bd=0,
                                highlightthickness=0,activestyle="none")
        self.listbox.pack(fill="both",expand=True,padx=8,pady=8)
        sb.config(command=self.listbox.yview)
        self.listbox.bind("<Delete>",self.remove_sel)
        tk.Label(left,text="💡 選中後按 Delete 可移除",font=("Segoe UI",9),
                 bg=BG,fg="#4A6080").pack(anchor="w",pady=4)

        # ══ 右側：設定面板（可捲動）════════════════
        right_wrap=tk.Frame(main,bg=BG,width=360)
        right_wrap.pack(side="right",fill="y",padx=(14,0))
        right_wrap.pack_propagate(False)
        right_canvas=tk.Canvas(right_wrap,bg=BG,highlightthickness=0,width=340)
        right_sb=tk.Scrollbar(right_wrap,orient="vertical",command=right_canvas.yview)
        right_canvas.configure(yscrollcommand=right_sb.set)
        right_sb.pack(side="right",fill="y")
        right_canvas.pack(side="left",fill="both",expand=True)
        right=tk.Frame(right_canvas,bg=BG)
        right_id=right_canvas.create_window((0,0),window=right,anchor="nw",width=340)
        def _update_scroll(e=None):
            right_canvas.configure(scrollregion=right_canvas.bbox("all"))
        right.bind("<Configure>",_update_scroll)
        def _scroll(e): right_canvas.yview_scroll(-1*(e.delta//120),"units")
        right_canvas.bind("<MouseWheel>",_scroll)
        right_wrap.bind("<MouseWheel>",_scroll)

        # ── 輸出資料夾 ───────────────────────────
        self._section(right,"📁  輸出資料夾")
        out_card=tk.Frame(right,bg=CARD); out_card.pack(fill="x",pady=2)
        tk.Entry(out_card,textvariable=self.out_dir,bg=CARD,fg=TEXT,
                 insertbackground="white",relief="flat",
                 font=("Segoe UI",11)).pack(side="left",fill="x",expand=True,
                                            ipady=8,padx=(10,4),pady=5)
        tk.Button(out_card,text="瀏覽",font=("Segoe UI",11),
                  bg="#1E3A58",fg=TEXT2,relief="flat",cursor="hand2",
                  padx=12,pady=5,command=self.browse_dir).pack(side="right",padx=4,pady=4)

        self.count_lbl=tk.Label(right,text="尚未選擇檔案",
                                font=("Segoe UI",11),bg=BG,fg="#4A7090")
        self.count_lbl.pack(anchor="w",pady=(4,0))

        self._sep(right)

        # ── 文件類型（大型切換按鈕）─────────────
        self._section(right,"📋  文件類型")
        self.doc_type=tk.StringVar(value="other")
        self._type_btns={}

        type_frame=tk.Frame(right,bg=BG); type_frame.pack(fill="x",pady=4)
        type_defs=[
            ("shipment","📦 出貨順序","浮水印過濾・日期合併"),
            ("other",   "📄 其他",    "單純轉檔"),
        ]
        for val,lbl,tip in type_defs:
            card=tk.Frame(type_frame,bg=CARD,cursor="hand2")
            card.pack(side="left",fill="x",expand=True,padx=(0,6))
            btn=tk.Label(card,text=lbl,font=("Segoe UI",11,"bold"),
                         bg=CARD,fg=TEXT2,padx=10,pady=10,cursor="hand2")
            btn.pack(fill="x")
            sub=tk.Label(card,text=tip,font=("Segoe UI",10),
                         bg=CARD,fg="#4A6080",padx=10,pady=8)
            sub.pack(fill="x",anchor="w")
            self._type_btns[val]=(card,btn,sub)
            for w in (card,btn,sub):
                w.bind("<Button-1>",lambda e,v=val: self._select_doc_type(v))

        self._sep(right)

        # ── 進階設定（可展開）───────────────────
        self._adv_open=tk.BooleanVar(value=False)
        adv_hdr=tk.Frame(right,bg="#0D1F30",cursor="hand2")
        adv_hdr.pack(fill="x")
        self._adv_arrow=tk.Label(adv_hdr,text="▶  進階設定",
                                  font=("Segoe UI",12,"bold"),
                                  bg="#0D1F30",fg=TEXT2,cursor="hand2")
        self._adv_arrow.pack(side="left",padx=12,pady=8)
        tk.Label(adv_hdr,text="（選填）",font=("Segoe UI",10),
                 bg="#0D1F30",fg="#3A5070").pack(side="left")
        for w in (adv_hdr, self._adv_arrow):
            w.bind("<Button-1>",self._toggle_adv)

        adv=tk.Frame(right,bg="#0D1F30")
        self._adv_frame=adv

        # 解析模式（每個選項獨立一列，不擠在一起）
        tk.Label(adv,text="解析模式",font=("Segoe UI",10,"bold"),
                 bg="#0D1F30",fg=TEXT2).pack(anchor="w",padx=12,pady=10)
        self.mode_var=tk.StringVar(value="auto")
        mode_defs=[
            ("auto",   "自動",    "先試框線偵測，失敗改用文字座標"),
            ("table",  "表格偵測","適合有明顯框線的 PDF"),
            ("coords", "文字座標","適合無框線或特殊版面的 PDF"),
        ]
        for val,lbl,tip in mode_defs:
            row=tk.Frame(adv,bg="#0D1F30"); row.pack(fill="x",padx=12,pady=1)
            tk.Radiobutton(row,text=lbl,variable=self.mode_var,value=val,
                           bg="#0D1F30",fg=TEXT,selectcolor="#1A3050",
                           activebackground="#0D1F30",
                           font=("Segoe UI",11,"bold"),
                           command=self._on_mode,width=9,anchor="w").pack(side="left")
            tk.Label(row,text=tip,font=("Segoe UI",10),
                     bg="#0D1F30",fg="#4A6080").pack(side="left",padx=(4,0))

        def _row(parent, label, var, tip=""):
            tk.Label(parent,text=label,font=("Segoe UI",11),
                     bg="#0D1F30",fg=TEXT2).pack(anchor="w",padx=12,pady=8)
            tk.Entry(parent,textvariable=var,bg="#0A1828",fg=TEXT,
                     insertbackground="white",relief="flat",
                     font=("Segoe UI",11)).pack(fill="x",padx=12,ipady=6,pady=1)
            if tip:
                tk.Label(parent,text=tip,font=("Segoe UI",9),
                         bg="#0D1F30",fg="#3A5070").pack(anchor="w",padx=12,pady=2)

        self.bounds_var=tk.StringVar()
        self.bounds_frame=tk.Frame(adv,bg="#0D1F30")
        self.bounds_frame.pack(fill="x")
        _row(self.bounds_frame,"欄界 x 座標（逗號分隔，留空=自動偵測）",
             self.bounds_var,"例：0,116,182,250,361,443,513")
        self.bounds_frame.pack_forget()

        self.wm_var=tk.StringVar()
        _row(adv,"浮水印關鍵字（逗號分隔）",
             self.wm_var,"例：B242,允將,CONFIDENTIAL")

        self.anchor_var=tk.StringVar(value="0")
        _row(adv,"錨點欄（第幾欄為主鍵，0開始）",
             self.anchor_var,"空白的列→合併備註到上一筆")

        self.merge_range_var=tk.StringVar()
        self.merge_frame=tk.Frame(adv,bg="#0D1F30")
        self.merge_frame.pack(fill="x")
        _row(self.merge_frame,"合併欄位範圍（起-迄，如 3-4）",
             self.merge_range_var,"將指定欄合併為一欄")
        self.merge_frame.pack_forget()

        self.skip_dup=tk.BooleanVar(value=True)
        ck=tk.Checkbutton(adv,text="自動刪除重複標題列",
                          variable=self.skip_dup,
                          bg="#0D1F30",fg=TEXT,selectcolor="#1A3050",
                          activebackground="#0D1F30",
                          font=("Segoe UI",11))
        ck.pack(anchor="w",padx=12,pady=(10,12))

        self._sep(right)

        # ── 進度條 ───────────────────────────────
        prog_frame=tk.Frame(right,bg=BG); prog_frame.pack(fill="x",pady=8)
        tk.Label(prog_frame,text="轉換進度",font=("Segoe UI",10,"bold"),
                 bg=BG,fg=TEXT2).pack(anchor="w")
        self.pbar=ttk.Progressbar(prog_frame,style="P.Horizontal.TProgressbar",
                                   orient="horizontal",mode="determinate")
        self.pbar.pack(fill="x",pady=(4,2))
        self.pbar_lbl=tk.Label(prog_frame,text="",font=("Segoe UI",11),
                                bg=BG,fg=TEXT2)
        self.pbar_lbl.pack(anchor="w")

        # ── 操作按鈕 ─────────────────────────────
        bc=dict(relief="flat",cursor="hand2")
        tk.Button(right,text="🔍  預覽 / 調整欄位",
                  font=("Segoe UI",12),bg="#1E3A58",fg=TEXT,
                  pady=10,command=self.open_preview,**bc).pack(fill="x",pady=(0,6))
        self.btn_start=tk.Button(right,text="▶  開始批次轉換",
                                  font=("Segoe UI",15,"bold"),
                                  bg=ACCENT,fg="white",pady=14,
                                  command=self.start_convert,**bc)
        self.btn_start.pack(fill="x",pady=(0,6))
        tk.Button(right,text="📂  開啟輸出資料夾",
                  font=("Segoe UI",11),bg=BG2,fg=TEXT2,
                  pady=8,command=self.open_out,**bc).pack(fill="x",pady=(0,8))


        # Log
        logf=tk.Frame(self,bg=BG2,height=148); logf.pack(fill="x",padx=16,pady=10); logf.pack_propagate(False)
        tk.Label(logf,text="執行記錄",font=("Segoe UI",11,"bold"),bg=BG2,fg=TEXT2).pack(anchor="w",padx=10,pady=6)
        lsb=tk.Scrollbar(logf); lsb.pack(side="right",fill="y")
        self.log_box=tk.Text(logf,yscrollcommand=lsb.set,bg=BG2,fg=TEXT2,font=("Consolas",10),
                             relief="flat",bd=0,state="disabled",wrap="word",highlightthickness=0)
        self.log_box.pack(fill="both",expand=True,padx=10,pady=(0,6))
        lsb.config(command=self.log_box.yview)
        self.log_box.tag_config("ok",foreground=SUCCESS)
        self.log_box.tag_config("err",foreground=ERROR)
        self.log_box.tag_config("warn",foreground=WARNING)

        # 所有元件建立完後才設定初始選中
        self.after(0, lambda: self._select_doc_type("other"))

    def _build_compare_tab(self, parent):
        """比對頁籤"""
        self._cmp_file_a=tk.StringVar()
        self._cmp_file_b=tk.StringVar()
        self._cmp_table_a=None
        self._cmp_table_b=None
        self._cmp_header=[]
        self._cmp_col_vars=[]   # Checkbutton 變數
        self._cmp_results=None

        # ── 上半：檔案選擇 ──────────────────────
        top=tk.Frame(parent,bg=BG); top.pack(fill="x",padx=16,pady=(12,6))
        tk.Label(top,text="檔案比對",font=("Segoe UI",16,"bold"),
                 bg=BG,fg=TEXT).pack(anchor="w",pady=(0,8))

        files_row=tk.Frame(top,bg=BG); files_row.pack(fill="x")
        for i,(lbl,var_attr) in enumerate([("📄 檔案 A（舊版）","_cmp_file_a"),
                                            ("📄 檔案 B（新版）","_cmp_file_b")]):
            col=tk.Frame(files_row,bg=CARD); col.pack(side="left",fill="x",expand=True,padx=(0,8) if i==0 else 0)
            tk.Label(col,text=lbl,font=("Segoe UI",11,"bold"),
                     bg=CARD,fg=TEXT2,padx=12).pack(anchor="w",pady=(8,2))
            row=tk.Frame(col,bg=CARD); row.pack(fill="x",padx=8,pady=(0,8))
            var=getattr(self,var_attr)
            tk.Entry(row,textvariable=var,bg="#0A1828",fg=TEXT,
                     insertbackground="white",relief="flat",
                     font=("Segoe UI",10),state="readonly").pack(side="left",fill="x",expand=True,ipady=6)
            tk.Button(row,text="選擇",font=("Segoe UI",10),
                      bg=ACCENT,fg="white",relief="flat",cursor="hand2",
                      padx=10,pady=4,
                      command=lambda v=var_attr: self._cmp_pick_file(v)
                      ).pack(side="right",padx=(6,0))

        # ── 載入按鈕 ────────────────────────────
        load_row=tk.Frame(top,bg=BG); load_row.pack(fill="x",pady=(8,0))
        tk.Button(load_row,text="📂  載入兩份檔案，選擇比對欄位",
                  font=("Segoe UI",12,"bold"),
                  bg="#1E3A58",fg=TEXT,relief="flat",cursor="hand2",
                  pady=10,command=self._cmp_load).pack(side="left",padx=(0,10))
        self._cmp_status=tk.Label(load_row,text="請先選擇兩份檔案",
                                   font=("Segoe UI",11),bg=BG,fg=TEXT2)
        self._cmp_status.pack(side="left")

        tk.Frame(parent,bg="#2A3A50",height=1).pack(fill="x",padx=16,pady=6)

        # ── 中：欄位選擇 ────────────────────────
        mid=tk.Frame(parent,bg=BG); mid.pack(fill="x",padx=16,pady=(0,6))
        tk.Label(mid,text="選擇比對欄位",font=("Segoe UI",12,"bold"),
                 bg=BG,fg=TEXT).pack(anchor="w",pady=(0,6))
        self._cmp_col_frame=tk.Frame(mid,bg=BG)
        self._cmp_col_frame.pack(fill="x")
        tk.Label(self._cmp_col_frame,text="（載入檔案後顯示）",
                 font=("Segoe UI",11),bg=BG,fg="#4A6080").pack(anchor="w")

        tk.Frame(parent,bg="#2A3A50",height=1).pack(fill="x",padx=16,pady=6)

        # ── 下半：比對結果預覽 ──────────────────
        bot_top=tk.Frame(parent,bg=BG); bot_top.pack(fill="x",padx=16,pady=(0,4))
        tk.Label(bot_top,text="比對結果",font=("Segoe UI",12,"bold"),
                 bg=BG,fg=TEXT).pack(side="left")
        self._cmp_summary=tk.Label(bot_top,text="",font=("Segoe UI",11),
                                    bg=BG,fg=TEXT2)
        self._cmp_summary.pack(side="left",padx=(12,0))
        bc=dict(relief="flat",cursor="hand2")
        tk.Button(bot_top,text="▶  開始比對",font=("Segoe UI",12,"bold"),
                  bg=ACCENT,fg="white",pady=7,padx=16,
                  command=self._cmp_run,**bc).pack(side="right")
        tk.Button(bot_top,text="💾  匯出差異報告 Excel",
                  font=("Segoe UI",11),bg="#1E3A58",fg=TEXT,
                  pady=7,padx=12,command=self._cmp_export,**bc).pack(side="right",padx=(0,8))

        # 結果表格
        res_frame=tk.Frame(parent,bg=BG); res_frame.pack(fill="both",expand=True,padx=16,pady=(0,12))
        self._cmp_tree_cols=["列號","差異類型","差異說明"]
        self._cmp_tree=ttk.Treeview(res_frame,columns=self._cmp_tree_cols,
                                     show="headings",height=15)
        for c,w in zip(self._cmp_tree_cols,[60,80,900]):
            self._cmp_tree.heading(c,text=c)
            self._cmp_tree.column(c,width=w,anchor="w")
        cvsb=ttk.Scrollbar(res_frame,orient="vertical",command=self._cmp_tree.yview)
        chsb=ttk.Scrollbar(res_frame,orient="horizontal",command=self._cmp_tree.xview)
        self._cmp_tree.configure(yscrollcommand=cvsb.set,xscrollcommand=chsb.set)
        self._cmp_tree.grid(row=0,column=0,sticky="nsew")
        cvsb.grid(row=0,column=1,sticky="ns")
        chsb.grid(row=1,column=0,sticky="ew")
        res_frame.rowconfigure(0,weight=1); res_frame.columnconfigure(0,weight=1)
        self._cmp_tree.tag_configure("add",foreground="#1E8449")
        self._cmp_tree.tag_configure("delete",foreground="#C0392B")
        self._cmp_tree.tag_configure("modify",foreground="#B7950B")
        self._cmp_tree.tag_configure("match",foreground="#4A6080")

    # ── 比對頁籤操作函式 ─────────────────────────
    def _cmp_pick_file(self, var_attr):
        f=filedialog.askopenfilename(
            title="選擇檔案",
            filetypes=[("PDF/Excel","*.pdf *.xlsx *.xls"),("所有","*.*")])
        if f: getattr(self,var_attr).set(f)

    def _cmp_load(self):
        fa=self._cmp_file_a.get(); fb=self._cmp_file_b.get()
        if not fa or not fb:
            messagebox.showwarning("提示","請先選擇兩份檔案！"); return
        self._cmp_status.config(text="載入中...")
        def _load():
            try:
                s=self._get_settings()
                ta=load_table_from_file(fa,s)
                tb=load_table_from_file(fb,s)
                self._cmp_table_a=ta
                self._cmp_table_b=tb
                header=ta[0] if ta else (tb[0] if tb else [])
                self._cmp_header=header
                self.after(0,lambda:self._cmp_show_cols(header))
                self.after(0,lambda:self._cmp_status.config(
                    text=f"✅ 已載入  A：{len(ta)-1}列  B：{len(tb)-1}列"))
            except Exception as e:
                self.after(0,lambda:self._cmp_status.config(text=f"❌ {e}"))
        threading.Thread(target=_load,daemon=True).start()

    def _cmp_show_cols(self, header):
        for w in self._cmp_col_frame.winfo_children(): w.destroy()
        self._cmp_col_vars=[]
        if not header:
            tk.Label(self._cmp_col_frame,text="找不到標題列",
                     font=("Segoe UI",11),bg=BG,fg=ERROR).pack(anchor="w"); return
        tk.Label(self._cmp_col_frame,
                 text="勾選要比對的欄位（未勾選的欄位不納入比對）：",
                 font=("Segoe UI",11),bg=BG,fg=TEXT2).pack(anchor="w",pady=(0,6))
        grid=tk.Frame(self._cmp_col_frame,bg=BG); grid.pack(fill="x")
        for ci,h in enumerate(header):
            v=tk.BooleanVar(value=False)
            self._cmp_col_vars.append(v)
            row,col=divmod(ci,4)
            cb=tk.Checkbutton(grid,text=f"  {h or f'欄{ci+1}'}",
                               variable=v,bg=BG,fg=TEXT,
                               selectcolor=CARD,activebackground=BG,
                               font=("Segoe UI",11))
            cb.grid(row=row,column=col,sticky="w",padx=(0,16),pady=2)

    def _cmp_run(self):
        if not self._cmp_table_a or not self._cmp_table_b:
            messagebox.showwarning("提示","請先載入兩份檔案！"); return
        cols=[i for i,v in enumerate(self._cmp_col_vars) if v.get()]
        if not cols:
            messagebox.showwarning("提示","請至少勾選一個比對欄位！"); return
        results,header=compare_tables(self._cmp_table_a,self._cmp_table_b,cols)
        self._cmp_results=results
        self._cmp_compare_cols=cols
        # 顯示結果
        tree=self._cmp_tree
        tree.delete(*tree.get_children())
        cnt={"add":0,"delete":0,"modify":0,"match":0}
        type_map={"add":"新增","delete":"刪除","modify":"修改","match":"相同"}
        for res in results:
            t=res["type"]; cnt[t]+=1
            if t=="match": continue
            idx=res["idx"]+2
            row_a=res["row_a"] or []; row_b=res["row_b"] or []
            if t=="add":
                desc=f"檔案B第{idx}列新增："+", ".join(
                    f"{header[c] if c<len(header) else f'欄{c+1}'}=「{clean(row_b[c]) if c<len(row_b) else ''}」"
                    for c in cols)
            elif t=="delete":
                desc=f"第{idx}列已從檔案B刪除："+", ".join(
                    f"{header[c] if c<len(header) else f'欄{c+1}'}=「{clean(row_a[c]) if c<len(row_a) else ''}」"
                    for c in cols)
            else:
                parts=[f"【{cn}】「{va}」→「{vb}」" for _,cn,va,vb in res["diffs"]]
                desc="  |  ".join(parts)
            tree.insert("","end",values=(idx,type_map[t],desc),tags=(t,))
        total=cnt["add"]+cnt["delete"]+cnt["modify"]
        self._cmp_summary.config(
            text=f"共 {total} 筆差異  新增 {cnt['add']}  刪除 {cnt['delete']}  修改 {cnt['modify']}  相同 {cnt['match']}",
            fg=ACCENT if total>0 else SUCCESS)

    def _cmp_export(self):
        if not self._cmp_results:
            messagebox.showwarning("提示","請先執行比對！"); return
        out=filedialog.asksaveasfilename(
            title="儲存差異報告",defaultextension=".xlsx",
            initialfile="差異報告.xlsx",filetypes=[("Excel","*.xlsx")])
        if not out: return
        try:
            write_diff_excel(self._cmp_results,self._cmp_header,
                             self._cmp_compare_cols,out)
            messagebox.showinfo("完成",f"已儲存：\n{out}")
            os.startfile(os.path.dirname(out))
        except Exception as e:
            messagebox.showerror("錯誤",str(e))

    def _on_mode(self):
        m=self.mode_var.get()
        if m=="coords":
            self.bounds_frame.pack(fill="x")
            self.merge_frame.pack(fill="x")
        else:
            self.bounds_frame.pack_forget()
            self.merge_frame.pack_forget()

    def _select_doc_type(self, val):
        """切換文件類型：更新視覺狀態 + 套用設定"""
        self.doc_type.set(val)
        for v,(card,btn,sub) in self._type_btns.items():
            if v==val:
                # 選中：亮色背景 + 白字
                card.config(bg=ACCENT if v=="shipment" else "#2A4A6A")
                btn.config(bg=ACCENT if v=="shipment" else "#2A4A6A",fg="white")
                sub.config(bg=ACCENT if v=="shipment" else "#2A4A6A",fg="#FFE0E8" if v=="shipment" else "#A0C8E0")
            else:
                # 未選中：暗色背景 + 灰字
                card.config(bg=CARD); btn.config(bg=CARD,fg=TEXT2); sub.config(bg=CARD,fg="#4A6080")
        self._on_doc_type()

    def _on_doc_type(self):
        """套用文件類型對應的預設設定"""
        # 尚未完成建構時不執行（避免存取未建立的屬性）
        if not hasattr(self,"wm_var") or not hasattr(self,"skip_dup"):
            return
        t=self.doc_type.get()
        if t=="shipment":
            self.skip_dup.set(True)
        else:
            self.wm_var.set("")
            self.skip_dup.set(True)

    def _toggle_adv(self, event=None):
        """展開/收合進階設定"""
        if self._adv_open.get():
            self._adv_frame.pack_forget()
            self._adv_arrow.config(text="▶  進階設定")
            self._adv_open.set(False)
        else:
            self._adv_frame.pack(fill="x",pady=(0,8))
            self._adv_arrow.config(text="▼  進階設定")
            self._adv_open.set(True)

    def _get_settings(self):
        doc_type=self.doc_type.get()
        base={
            "mode": self.mode_var.get(),
            "watermark_keywords": self.wm_var.get(),
            "anchor_col": self.anchor_var.get().strip() or "0",
            "skip_duplicate_header": self.skip_dup.get(),
            "col_bounds": self.bounds_var.get(),
            "merge_col_range": self.merge_range_var.get(),
            "doc_type": doc_type,
        }
        if doc_type=="other":
            # 其他模式：關閉日期填充和合併（忽略 watermark_keywords 設定）
            base["watermark_keywords"]=""
            base["skip_duplicate_header"]=False
            base["fill_dates"]=False
            base["merge_dates"]=False
        else:
            # 出貨順序：啟用所有處理
            base["fill_dates"]=True
            base["merge_dates"]=True
        return base

    def add_files(self):
        fs=filedialog.askopenfilenames(title="選擇 PDF",filetypes=[("PDF","*.pdf"),("所有","*.*")])
        for f in fs:
            if f not in self.pdf_files:
                self.pdf_files.append(f)
                self.listbox.insert("end",os.path.basename(f))
                # 自動設定輸出路徑為第一個檔案的資料夾
                if not self.out_dir.get():
                    self.out_dir.set(os.path.dirname(f))
        n=len(self.pdf_files)
        self.count_lbl.config(text=f"已選 {n} 個 PDF" if n else "尚未選擇檔案")

    def remove_sel(self,e=None):
        for i in reversed(self.listbox.curselection()): self.listbox.delete(i); self.pdf_files.pop(i)

    def clear_files(self):
        self.pdf_files.clear(); self.listbox.delete(0,"end")
        self.count_lbl.config(text="尚未選擇檔案")
    def browse_dir(self):
        d=filedialog.askdirectory()
        if d: self.out_dir.set(d)
    def open_out(self):
        d=self.out_dir.get()
        if os.path.isdir(d): os.startfile(d)

    def log(self,msg,tag="info"):
        self.log_box.config(state="normal"); self.log_box.insert("end",msg+"\n",tag)
        self.log_box.see("end"); self.log_box.config(state="disabled")

    def open_preview(self):
        if not self.pdf_files: messagebox.showwarning("提示","請先新增 PDF！"); return
        sel=self.listbox.curselection()
        pdf=self.pdf_files[sel[0] if sel else 0]
        settings=self._get_settings()
        self.log(f"載入預覽：{os.path.basename(pdf)}")
        def _load():
            try:
                table,strategy=convert(pdf,None,settings,log_cb=lambda m:self.after(0,lambda m=m:self.log(m)))
                self.after(0,lambda:PreviewWin(self,pdf,table,strategy,settings))
            except Exception as e:
                self.after(0,lambda:self.log(f"預覽失敗：{e}","err"))
        threading.Thread(target=_load,daemon=True).start()

    def start_convert(self):
        if self.running: return
        if not self.pdf_files: messagebox.showwarning("提示","請先新增 PDF！"); return
        if not os.path.isdir(self.out_dir.get()): messagebox.showerror("錯誤",f"資料夾不存在：{self.out_dir.get()}"); return
        self.running=True; self.btn_start.config(state="disabled",text="⏳  轉換中...")
        self.pbar["value"]=0; threading.Thread(target=self._run,daemon=True).start()

    def _run(self):
        files=list(self.pdf_files); out_dir=self.out_dir.get()
        settings=self._get_settings(); ok=fail=0
        self.log("═"*44); self.log(f"開始轉換 {len(files)} 個（模式：{settings['mode']}）")
        for i,pdf in enumerate(files):
            fname=os.path.basename(pdf); self.log(f"\n[{i+1}/{len(files)}] {fname}")
            self.after(0,lambda v=i/len(files)*100:self._set_pbar(v))
            out=os.path.join(out_dir,os.path.splitext(fname)[0]+".xlsx")
            if os.path.exists(out):
                ans=messagebox.askyesno("檔案已存在",
                    f"檔案已存在，是否覆蓋？\n{os.path.basename(out)}")
                if not ans:
                    self.log(f"  ⏭  跳過（使用者取消）","warn"); continue
            try:
                table,strategy=convert(pdf,out,settings,log_cb=lambda m:self.after(0,lambda m=m:self.log(m)))
                ok+=1; self.log(f"  ✅ {max(0,len(table)-1)} 筆 [{strategy}] → {os.path.basename(out)}","ok")
            except Exception as e:
                fail+=1; self.log(f"  ❌ {e}","err")
        self.after(0,lambda:self._set_pbar(100))
        self.log("\n"+"═"*44)
        self.log(f"完成：成功 {ok} / 失敗 {fail}","ok" if fail==0 else "err")
        self.after(0,self._done)

    def _set_pbar(self,v): self.pbar["value"]=v; self.pbar_lbl.config(text=f"{v:.0f}%")
    def _done(self): self.running=False; self.btn_start.config(state="normal",text="▶  開始批次轉換")


class PreviewWin(tk.Toplevel):
    def __init__(self,master,pdf_path,table,strategy,settings):
        super().__init__(master)
        self.pdf_path=pdf_path; self.table=table
        self.strategy=strategy; self.settings=settings; self.col_vars=[]
        self.title(f"預覽 — {os.path.basename(pdf_path)}")
        self.geometry("1300x740"); self.configure(bg=BG); self._build()

    def _build(self):
        info=tk.Frame(self,bg=BG2,height=32); info.pack(fill="x"); info.pack_propagate(False)
        nrows=len(self.table); ncols=max((len(r) for r in self.table),default=0)
        tk.Label(info,text=f"  策略：{self.strategy}    {nrows} 列 × {ncols} 欄    （前 50 列預覽）",
                 font=("Segoe UI",9),bg=BG2,fg=TEXT2).pack(side="left",padx=8)

        if self.table:
            hf=tk.Frame(self,bg=BG); hf.pack(fill="x",padx=10,pady=6)
            tk.Label(hf,text="欄位名稱（可修改後點「套用」更新預覽）：",font=("Segoe UI",9,"bold"),bg=BG,fg=TEXT).pack(side="left")
            ce=tk.Frame(self,bg=BG); ce.pack(fill="x",padx=10,pady=4)
            for ci,hval in enumerate(self.table[0]):
                v=tk.StringVar(value=clean(hval)); self.col_vars.append(v)
                f=tk.Frame(ce,bg=CARD,padx=4,pady=2); f.pack(side="left",padx=(0,3))
                tk.Label(f,text=f"欄{ci+1}",font=("Segoe UI",7),bg=CARD,fg=TEXT2).pack()
                tk.Entry(f,textvariable=v,width=12,bg=BG2,fg=TEXT,insertbackground="white",relief="flat",font=("Segoe UI",10)).pack(ipady=3)

        tf=tk.Frame(self,bg=BG); tf.pack(fill="both",expand=True,padx=10,pady=4)
        preview=self.table[:50]
        ncols2=max((len(r) for r in preview),default=1)
        cols=[f"C{i+1}" for i in range(ncols2)]
        self.tree=ttk.Treeview(tf,columns=cols,show="headings",height=17)
        cw=max(70,min(160,900//max(len(cols),1)))
        for c in cols: self.tree.heading(c,text=c); self.tree.column(c,width=cw,anchor="w")
        vsb=ttk.Scrollbar(tf,orient="vertical",command=self.tree.yview)
        hsb=ttk.Scrollbar(tf,orient="horizontal",command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,xscrollcommand=hsb.set)
        self.tree.grid(row=0,column=0,sticky="nsew"); vsb.grid(row=0,column=1,sticky="ns"); hsb.grid(row=1,column=0,sticky="ew")
        tf.rowconfigure(0,weight=1); tf.columnconfigure(0,weight=1)
        # Style 已在 App.__init__ 設定，這裡不重複初始化
        self._fill(preview)
        self.tree.bind("<Double-1>",self._on_double_click)
        tk.Label(tf,text="💡 雙擊儲存格可直接編輯　Enter 換行　Ctrl+Enter / Tab 儲存　Esc 取消",font=("Segoe UI",10),bg=BG,fg=TEXT2).grid(row=2,column=0,sticky="w",pady=4)

        bf=tk.Frame(self,bg=BG); bf.pack(fill="x",padx=10,pady=8)
        tk.Button(bf,text="🔄  套用欄位名稱",font=("Segoe UI",11),bg=CARD,fg=TEXT,relief="flat",cursor="hand2",padx=14,pady=6,command=self._apply).pack(side="left",padx=(0,8))
        tk.Button(bf,text="💾  另存為 Excel",font=("Segoe UI",12,"bold"),bg=ACCENT,fg="white",relief="flat",cursor="hand2",padx=18,pady=7,command=self.save).pack(side="right")
        tk.Button(bf,text="關閉",font=("Segoe UI",11),bg=BG2,fg=TEXT2,relief="flat",cursor="hand2",padx=14,pady=6,command=self.destroy).pack(side="right",padx=(0,6))

    def _fill(self,data):
        self.tree.delete(*self.tree.get_children())
        if not data: return
        header=[v.get() if i<len(self.col_vars) else f"欄{i+1}" for i,v in enumerate(self.col_vars)] if self.col_vars else [clean(v) or f"欄{i+1}" for i,v in enumerate(data[0])]
        ncols=max(len(r) for r in data)
        cols=[f"C{i+1}" for i in range(ncols)]
        for c,h in zip(cols,header): self.tree.heading(c,text=h)
        for row in data[1:]:
            vals=[row[i] if i<len(row) else "" for i in range(ncols)]
            self.tree.insert("","end",values=[str(v).replace("\n"," / ") for v in vals])

    def _apply(self): self._fill(self.table[:50])

    def _on_double_click(self, event):
        """雙擊儲存格 → inline Text 編輯
        Enter = 換行   Ctrl+Enter / Tab = 儲存   Esc = 取消
        """
        item=self.tree.focus()
        if not item: return
        col_id=self.tree.identify_column(event.x)
        ci=int(col_id.replace("#",""))-1
        items=list(self.tree.get_children())
        if item not in items: return
        ri=items.index(item)+1
        if ri>=len(self.table): return
        cur_val=self.table[ri][ci] if ci<len(self.table[ri]) else ""
        bbox=self.tree.bbox(item, col_id)
        if not bbox: return
        x,y,w,h=bbox
        # 統一用 Text 框，高度依內容自動設定（最少2行，最多8行）
        n_lines=max(cur_val.count("\n")+1, 2)
        edit_h=min(n_lines, 8)
        widget=tk.Text(self.tree, bg="#FFFDE7", fg="black",
                       relief="solid", bd=1,
                       font=("Segoe UI",10), wrap="word",
                       height=edit_h, insertbackground="black",
                       selectbackground=ACCENT2)
        widget.insert("1.0", cur_val)
        widget.place(x=x, y=y, width=max(w,220), height=edit_h*18+4)
        widget.focus_set()
        widget.mark_set("insert","end")

        orig_val=cur_val  # 用於 Esc 取消

        def _commit(e=None):
            new_val=widget.get("1.0","end-1c")
            while len(self.table[ri])<ci+1: self.table[ri].append("")
            self.table[ri][ci]=new_val
            widget.destroy()
            self._fill(self.table[:50])

        def _cancel(e=None):
            widget.destroy()

        def _key(e):
            # Ctrl+Enter 或 Tab → 儲存
            if e.keysym in ("Tab",) or (e.keysym=="Return" and (e.state&0x4)):
                _commit(); return "break"
            # Esc → 取消
            if e.keysym=="Escape":
                _cancel(); return "break"
            # 其他鍵正常處理（Enter 換行）

        widget.bind("<KeyPress>", _key)
        widget.bind("<FocusOut>", _commit)

    def save(self):
        out=filedialog.asksaveasfilename(title="另存 Excel",defaultextension=".xlsx",
            initialfile=os.path.splitext(os.path.basename(self.pdf_path))[0]+".xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not out: return
        try:
            col_map={i:v.get() for i,v in enumerate(self.col_vars) if v.get()} if self.col_vars else None
            mi=detect_merges(self.table)
            write_excel(self.table,mi,out,col_map=col_map)
            messagebox.showinfo("完成",f"已儲存：\n{out}")
            os.startfile(os.path.dirname(out))
        except Exception as e:
            messagebox.showerror("錯誤",str(e))

if __name__=="__main__":
    App().mainloop()
